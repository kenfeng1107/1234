from openpyxl import load_workbook, Workbook
from datetime import datetime

# 取得當前時間
IMA = datetime.today()

YYYYYY = IMA.year   # 年
MMMMMM = IMA.month  # 月
DDDDDD = IMA.day    # 日

# 組合年、月、日作為新工作表名稱
daTedaTe = [YYYYYY, MMMMMM, DDDDDD]
DATAcreate = '-'.join(map(str, daTedaTe))

# 學生姓名與 RFID 卡號的對應
rfid_to_name = {
    "3801031571": "1.A",
    "1032025142": "2.B",
    "1030886662": "3.C",
    "1033857606": "4.D",
}

# 嘗試載入 Excel 文件
try:
    Archieve = load_workbook('nega.xlsx')
except FileNotFoundError:
    Archieve = Workbook()

# 檢查是否已有與當前日期匹配的工作表名稱
if DATAcreate not in Archieve.sheetnames:
    sheet = Archieve.create_sheet(DATAcreate)  # 創建新工作表
    
    # 添加標題
    sheet['A1'] = '姓名/點名表'
    
    # 填寫第一到第八節課
    lessons = ['第一節課', '第二節課', '第三節課', '第四節課', '第五節課', '第六節課', '第七節課', '第八節課']
    for i, lesson in enumerate(lessons, start=2):  # 從 B1 開始填入
        sheet.cell(row=1, column=i, value=lesson)
    
    # 添加學生姓名列
    for i, (rfid, name) in enumerate(rfid_to_name.items(), start=2):
        sheet.cell(row=i, column=1, value=name)  # 在A列填入學生姓名

    # 設置列寬和行高
    sheet.column_dimensions['A'].width = 20  # 學生姓名列寬
    for col_num in range(2, len(lessons) + 2):  # 設置課程列寬
        sheet.column_dimensions[chr(64 + col_num)].width = 15  # 寬度適中

    for row_num in range(1, len(rfid_to_name) + 2):  # 設置行高
        sheet.row_dimensions[row_num].height = 20

    # 保存文件
    Archieve.save('nega.xlsx')
    print(f"新工作表 {DATAcreate} 已創建並保存。")
else:
    print(f"工作表 {DATAcreate} 已存在。")

# 點名記錄部分
def x(rfid_input):
    IMA = datetime.today()
    HHHHHH = IMA.hour   # 小時
    MIMIMI = IMA.minute  # 分鐘
    PPHHHH = HHHHHH * 60 + MIMIMI  # 小時換算為分鐘

    # 定義時間範圍和規則
    time_ranges = [
        (480, 530, "第一節課"),  # 08:00-08:50
        (550, 600, "第二節課"),  # 09:10-10:00
        (610, 670, "第三節課"),  # 10:10-11:00
        (680, 730, "第四節課"),  # 11:10-12:00
        (780, 830, "第五節課"),  # 13:00-13:50
        (840, 890, "第六節課"),  # 14:00-14:50
        (900, 950, "第七節課"),  # 15:00-15:50
        (970, 1020, "第八節課")  # 16:10-17:00
    ]

    def get_attendance_status(minute):
        for start, end, lesson in time_ranges:
            # 通過調整對比的方式，更精確判斷
            if start <= minute <= end:
                elapsed = minute - start
                if elapsed <= 5:
                    return lesson, "正常", IMA.strftime('%H:%M:%S')  # 正常記錄時間
                elif elapsed <= 15:
                    return lesson, "遲到", None  # 顯示遲到但不記錄時間
                else:
                    return lesson, "曠課", None  # 顯示曠課但不記錄時間
        return None, None, None  # 不在任何時間範圍

    def record_attendance():
        # 嘗試載入現有工作簿
        workbook = load_workbook('nega.xlsx')

        # 獲取當前工作表
        sheet = workbook[DATAcreate]  # 獲取對應日期的工作表
        lesson, status, current_time = get_attendance_status(PPHHHH)

        # 根據 RFID 卡號識別學生
        student_name = rfid_to_name.get(rfid_input, "未知學生")

        # 查找學生在表格中的行
        row = None
        for i in range(2, len(rfid_to_name) + 2):  # 從第2行開始找
            if sheet.cell(row=i, column=1).value == student_name:
                row = i
                break

        if row and lesson:
            # 找到課程列
            lesson_columns = {
                "第一節課": 2,
                "第二節課": 3,
                "第三節課": 4,
                "第四節課": 5,
                "第五節課": 6,
                "第六節課": 7,
                "第七節課": 8,
                "第八節課": 9,
            }
            column = lesson_columns[lesson]

            if current_time:  # 記錄時間
                sheet.cell(row=row, column=column, value=current_time)
            else:  # 記錄 "遲到" 或 "曠課"
                sheet.cell(row=row, column=column, value=status)

            print(f"記錄成功：{student_name} - {lesson} - {status}")
        else:
            print(f"記錄失敗：未找到匹配節次或學生信息錯誤。")

        # 保存文件
        workbook.save('nega.xlsx')

    # 記錄考勤
    record_attendance()

# 模擬從 RFID 讀取輸入（在實際中，這部分將來自 RFID 讀取器）
rfid_input = input("請輸入 RFID (或按 Enter 結束): ")
if rfid_input:
    x(rfid_input)
    print(f"已記錄: RFID={rfid_input}, 學生={rfid_to_name.get(rfid_input, '未知學生')}, 時間={IMA.strftime('%Y-%m-%d %H:%M:%S')}")
else:
    print("沒有輸入 RFID，程序結束。")